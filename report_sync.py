#!/usr/bin/env python3
"""
report_sync.py — 作業報告書 Excel ↔ kintai データ変換ライブラリ

Work_Report フォルダの xlsx を読み書きし、PWA(kintai)データと相互変換する。
jinjer_server.py から import して使う。

単体テスト:
  python3 report_sync.py
"""
import json
import os
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill
    _OPENPYXL_OK = True
except ImportError:
    openpyxl = None          # type: ignore
    PatternFill = None       # type: ignore
    _OPENPYXL_OK = False


def _require_openpyxl():
    if not _OPENPYXL_OK:
        raise ImportError('openpyxl が必要です: pip install openpyxl')

try:
    import urllib.request
    _REQUESTS_OK = True
except Exception:
    _REQUESTS_OK = False

# ===== パス設定 =====
WORK_REPORT_DIR = Path.home() / 'Library/Mobile Documents/com~apple~CloudDocs/:root/Work_Report'
WEEKDAY_JA = ['月', '火', '水', '木', '金', '土', '日']

# Excel 列インデックス (0 始まり)
COL_DATE    = 0
COL_WEEKDAY = 1
COL_START   = 2
COL_END     = 3
COL_BREAK   = 4
COL_TOTAL   = 5
COL_CONTENT = 6
COL_NOTE    = 11   # 備考（在宅/出社/祝日など）

# メタ行 (0-indexed row, col)
META_CLIENT  = (1, 2)   # 取引先名
META_PROJECT = (1, 10)  # 案件名
META_DEPT    = (2, 2)   # 部署名
META_PLACE   = (2, 10)  # 作業場所
META_TECH    = (6, 10)  # 技術者名
META_START   = (4, 10)  # 作業開始日
META_END     = (5, 10)  # 作業終了日
META_DEFAULT_START = (6, 2)   # 基本就業 開始
META_DEFAULT_END   = (6, 3)   # 基本就業 終了
META_DEFAULT_BREAK = (6, 4)   # 基本就業 休憩

DATA_START_ROW = 12  # 0-indexed でのデータ開始行（行13 = index 12）


# ===== ステータス判定 =====

def detect_status(filename: str) -> str:
    """ファイル名から承認ステータスを判定する"""
    if '上長承認済' in filename or '承認済' in filename:
        return 'approved'
    if '押印済' in filename:
        return 'approved'
    if '確認' in filename:
        return 'reviewing'
    return 'draft'


def detect_month_from_filename(filename: str) -> tuple[str, str] | None:
    """ファイル名から (year, month) を抽出する。例: '202602分_...' → ('2026', '02')"""
    m = re.match(r'^(\d{4})(\d{2})分', filename)
    if m:
        return m.group(1), m.group(2)
    return None


# ===== 時刻変換ユーティリティ =====

def time_to_str(val) -> str:
    """datetime.time または datetime.timedelta → 'HH:MM' 文字列"""
    if val is None:
        return ''
    if hasattr(val, 'hour'):  # datetime.time
        return f'{val.hour:02d}:{val.minute:02d}'
    if isinstance(val, timedelta):
        total = int(val.total_seconds())
        h = total // 3600
        m = (total % 3600) // 60
        return f'{h:02d}:{m:02d}'
    return str(val)


def str_to_time(s: str):
    """'HH:MM' → datetime.time"""
    from datetime import time as dtime
    if not s:
        return None
    try:
        h, m = map(int, s.split(':'))
        return dtime(h, m)
    except Exception:
        return None


def date_to_str(val) -> str:
    """datetime.datetime → 'YYYY-MM-DD'"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    return str(val)


# ===== ファイル一覧 =====

def list_reports() -> list[dict]:
    """Work_Report フォルダのファイル一覧を返す"""
    if not WORK_REPORT_DIR.exists():
        return []

    reports = []
    for f in sorted(WORK_REPORT_DIR.iterdir()):
        if not f.name.endswith('.xlsx') or f.name.startswith('~$'):
            continue
        ym = detect_month_from_filename(f.name)
        if not ym:
            continue
        year, month = ym
        reports.append({
            'filename': f.name,
            'year': year,
            'month': month,
            'status': detect_status(f.name),
            'size': f.stat().st_size,
            'modified': datetime.fromtimestamp(f.stat().st_mtime).strftime('%Y-%m-%d %H:%M'),
        })
    return reports


# ===== Excel 読み込み =====

def read_report(year: str, month: str) -> dict | None:
    """指定年月の作業報告書 Excel を読み込んで dict に変換する"""
    _require_openpyxl()
    target = f'{year}{month}'
    found = None
    if not WORK_REPORT_DIR.exists():
        return None

    for f in WORK_REPORT_DIR.iterdir():
        if f.name.endswith('.xlsx') and f.name.startswith(target) and not f.name.startswith('~$'):
            found = f
            break

    if not found:
        return None

    wb = openpyxl.load_workbook(str(found), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, c):
        try:
            return rows[r][c]
        except IndexError:
            return None

    # メタ情報
    meta = {
        'client':  cell(*META_CLIENT)  or '',
        'project': cell(*META_PROJECT) or '',
        'dept':    cell(*META_DEPT)    or '',
        'place':   cell(*META_PLACE)   or '',
        'tech':    cell(*META_TECH)    or '',
        'start_date': date_to_str(cell(*META_START)),
        'end_date':   date_to_str(cell(*META_END)),
        'default_start': time_to_str(cell(*META_DEFAULT_START)),
        'default_end':   time_to_str(cell(*META_DEFAULT_END)),
        'default_break': time_to_str(cell(*META_DEFAULT_BREAK)),
    }

    # 日別データ
    # テンプレートによっては日付セルが数式の場合 data_only=True で None になるため、
    # 行インデックスと作業開始日から日付を推定するフォールバックを使用
    try:
        report_start = datetime.strptime(meta['start_date'], '%Y-%m-%d') if meta['start_date'] else None
    except Exception:
        report_start = None

    days = []
    data_day_idx = 0  # 有効な日付行のカウンタ（start_dateフォールバック用）
    for row in rows[DATA_START_ROW:]:
        date_val = row[COL_DATE] if len(row) > COL_DATE else None

        if date_val is None and report_start is not None:
            # 数式セルで計算結果がない場合: 行番号から日付を推定
            import calendar
            days_in_m = calendar.monthrange(report_start.year, report_start.month)[1]
            if data_day_idx >= days_in_m:
                break
            inferred_date = datetime(report_start.year, report_start.month, data_day_idx + 1)
            d = date_to_str(inferred_date)
            weekday_str = WEEKDAY_JA[inferred_date.weekday()]
        elif not isinstance(date_val, (datetime, date)):
            # 合計行など非日付行はスキップ
            break
        else:
            d = date_to_str(date_val)
            weekday_str = str(row[COL_WEEKDAY] or '') if len(row) > COL_WEEKDAY else ''

        if not d:
            break

        data_day_idx += 1

        content = row[COL_CONTENT] if len(row) > COL_CONTENT else None
        note    = row[COL_NOTE]    if len(row) > COL_NOTE    else None

        days.append({
            'date':     d,
            'weekday':  weekday_str,
            'start':    time_to_str(row[COL_START] if len(row) > COL_START else None),
            'end':      time_to_str(row[COL_END]   if len(row) > COL_END   else None),
            'break':    time_to_str(row[COL_BREAK] if len(row) > COL_BREAK else None),
            'total':    time_to_str(row[COL_TOTAL] if len(row) > COL_TOTAL else None),
            'content':  str(content) if content else '',
            'note':     str(note)    if note    else '',
        })

    return {
        'filename': found.name,
        'status':   detect_status(found.name),
        'meta':     meta,
        'days':     days,
    }


# ===== kintai データ → Excel 書き込み =====

def write_report_from_kintai(year: str, month: str, kintai_data: dict) -> dict:
    """
    kintai の月データを作業報告書 Excel に書き込む。

    kintai_data 形式:
    {
      "YYYY-MM-DD": {
        "status": "在宅" | "出社" | "休み" | "休日" | "祝日" | "未",
        "start": "09:00",
        "end": "18:00",
        "memo": "作業内容テキスト"
      }, ...
    }
    戻り値: {"ok": True, "path": "...", "updated": N}
    """
    _require_openpyxl()
    target = f'{year}{month}'
    found = None
    if not WORK_REPORT_DIR.exists():
        return {'ok': False, 'error': 'Work_Report ディレクトリが見つかりません'}

    for f in WORK_REPORT_DIR.iterdir():
        if f.name.endswith('.xlsx') and f.name.startswith(target) and not f.name.startswith('~$'):
            found = f
            break

    if not found:
        return {'ok': False, 'error': f'{year}{month} の Excel ファイルが見つかりません'}

    wb = openpyxl.load_workbook(str(found))
    ws = wb.active
    rows_in_sheet = list(ws.iter_rows())

    # ステータス → 備考文字列
    status_to_note = {
        '在宅': '在宅',
        '出社': '出社',
        '休み': '有給',
        '祝日': '祝日',
        '休日': '休日',
        '未':   '',
    }

    def safe_write(cell, val):
        try:
            cell.value = val
        except AttributeError:
            pass

    updated = 0
    for row_cells in rows_in_sheet[DATA_START_ROW:]:
        date_cell = row_cells[COL_DATE]
        if not isinstance(date_cell.value, (datetime, date)):
            continue

        date_key = date_to_str(date_cell.value)
        if date_key not in kintai_data:
            continue

        kd = kintai_data[date_key]
        status = kd.get('status', '未')
        start  = kd.get('start', '')
        end    = kd.get('end', '')
        memo   = kd.get('memo', '')

        # 稼働日（在宅/出社）の場合のみ時間を書き込む
        if status in ('在宅', '出社') and start and end:
            safe_write(row_cells[COL_START], str_to_time(start))
            safe_write(row_cells[COL_END],   str_to_time(end))

            # 休憩は既存値を維持（空なら基本値 01:00 を入力）
            if row_cells[COL_BREAK].value is None:
                from datetime import time as dtime
                safe_write(row_cells[COL_BREAK], dtime(1, 0))

            # 合計時間を計算
            try:
                from datetime import time as dtime
                s_h, s_m = map(int, start.split(':'))
                e_h, e_m = map(int, end.split(':'))
                brk_val = row_cells[COL_BREAK].value
                if isinstance(brk_val, timedelta):
                    brk_h = int(brk_val.total_seconds() // 3600)
                    brk_m = int((brk_val.total_seconds() % 3600) // 60)
                elif hasattr(brk_val, 'hour'):
                    brk_h, brk_m = brk_val.hour, brk_val.minute
                else:
                    brk_h, brk_m = 1, 0
                total_min = (e_h * 60 + e_m) - (s_h * 60 + s_m) - (brk_h * 60 + brk_m)
                if total_min > 0:
                    row_cells[COL_TOTAL].value = dtime(total_min // 60, total_min % 60)
            except Exception:
                pass

        # 作業内容（メモがあれば）
        if memo:
            safe_write(row_cells[COL_CONTENT], memo)

        # 備考（在宅/出社/祝日など）
        note = status_to_note.get(status, '')
        if note:
            safe_write(row_cells[COL_NOTE], note)

        updated += 1

    wb.save(str(found))
    return {'ok': True, 'path': str(found), 'updated': updated}


# ===== 翌月 Excel 自動生成 =====

def fetch_holidays(year: int) -> dict[str, str]:
    """内閣府祝日 API から祝日情報を取得する"""
    url = f'https://holidays-jp.github.io/api/v1/{year}/date.json'
    try:
        with urllib.request.urlopen(url, timeout=10) as res:
            return json.loads(res.read().decode('utf-8'))
    except Exception:
        return {}


def create_next_month_report(year: str, month: str) -> dict:
    """
    指定年月の翌月作業報告書 Excel を自動生成する。
    前月または最新の Excel をテンプレートとしてコピーし、
    日付・曜日・祝日を自動入力する。

    year: '2026', month: '02' → 2026年3月分を生成
    戻り値: {"ok": True, "path": "...", "filename": "..."}
    """
    _require_openpyxl()
    # 翌月を計算
    y, m = int(year), int(month)
    m += 1
    if m > 12:
        m = 1
        y += 1
    next_year  = str(y)
    next_month = f'{m:02d}'
    next_ym    = f'{next_year}{next_month}'

    # 既存チェック
    if WORK_REPORT_DIR.exists():
        for f in WORK_REPORT_DIR.iterdir():
            if f.name.startswith(next_ym) and f.name.endswith('.xlsx'):
                return {'ok': False, 'error': f'{next_year}年{int(next_month)}月分は既に存在します: {f.name}'}

    # テンプレート検索（前月 → 最新ファイル）
    template = None
    cur_ym = f'{year}{month}'
    if WORK_REPORT_DIR.exists():
        for f in sorted(WORK_REPORT_DIR.iterdir(), reverse=True):
            if f.name.endswith('.xlsx') and not f.name.startswith('~$'):
                ym = detect_month_from_filename(f.name)
                if ym and f'{ym[0]}{ym[1]}' == cur_ym:
                    template = f
                    break
        if not template:
            for f in sorted(WORK_REPORT_DIR.iterdir(), reverse=True):
                if f.name.endswith('.xlsx') and not f.name.startswith('~$'):
                    template = f
                    break

    if not template:
        return {'ok': False, 'error': 'テンプレートとなる Excel が見つかりません'}

    # 翌月の日数・祝日を取得
    import calendar
    days_in_month = calendar.monthrange(y, m)[1]
    holidays = fetch_holidays(y)

    # テンプレートをコピー
    new_filename = f'{next_ym}分_作業報告書_(柳田侑佑).xlsx'
    new_path = WORK_REPORT_DIR / new_filename
    shutil.copy2(str(template), str(new_path))

    wb = openpyxl.load_workbook(str(new_path))
    ws = wb.active
    rows_in_sheet = list(ws.iter_rows())

    # 作業開始日・終了日を更新
    start_date = datetime(y, m, 1)
    end_date   = datetime(y, m, days_in_month)
    if len(rows_in_sheet) > META_START[0]:
        rows_in_sheet[META_START[0]][META_START[1]].value = start_date
        rows_in_sheet[META_END[0]][META_END[1]].value     = end_date

    # 作業場所のリセット
    if len(rows_in_sheet) > META_PLACE[0]:
        rows_in_sheet[META_PLACE[0]][META_PLACE[1]].value = '在宅'

    # 日別データ行をクリアして翌月の日付を入力
    # ※ テンプレートの日付セルは数式で自動計算されているため、
    #   行インデックスで判断し、値を直接上書きする
    data_rows = rows_in_sheet[DATA_START_ROW: DATA_START_ROW + 37]

    day_idx = 0
    for row_cells in data_rows:
        def safe_set(cell, val):
            try:
                cell.value = val
            except AttributeError:
                pass

        if day_idx >= days_in_month:
            # 余分な行をクリア
            safe_set(row_cells[COL_DATE],    None)
            safe_set(row_cells[COL_WEEKDAY], None)
            safe_set(row_cells[COL_START],   None)
            safe_set(row_cells[COL_END],     None)
            safe_set(row_cells[COL_BREAK],   None)
            safe_set(row_cells[COL_TOTAL],   None)
            safe_set(row_cells[COL_CONTENT], None)
            safe_set(row_cells[COL_NOTE],    None)
            day_idx += 1
            continue

        d = date(y, m, day_idx + 1)
        try:
            row_cells[COL_DATE].value    = datetime(y, m, day_idx + 1)
        except AttributeError:
            pass
        try:
            row_cells[COL_WEEKDAY].value = WEEKDAY_JA[d.weekday()]
        except AttributeError:
            pass

        # 土日・祝日の処理
        date_str = d.strftime('%Y-%m-%d')
        def safe_set(cell, val):
            """マージセル（MergedCell）はスキップ"""
            try:
                cell.value = val
            except AttributeError:
                pass

        if d.weekday() >= 5:  # 土日
            safe_set(row_cells[COL_START],   None)
            safe_set(row_cells[COL_END],     None)
            safe_set(row_cells[COL_BREAK],   None)
            safe_set(row_cells[COL_TOTAL],   None)
            safe_set(row_cells[COL_CONTENT], None)
            safe_set(row_cells[COL_NOTE],    None)
        elif date_str in holidays:
            safe_set(row_cells[COL_START],   None)
            safe_set(row_cells[COL_END],     None)
            safe_set(row_cells[COL_BREAK],   None)
            safe_set(row_cells[COL_TOTAL],   None)
            safe_set(row_cells[COL_CONTENT], holidays[date_str])
            safe_set(row_cells[COL_NOTE],    '祝日')
        else:
            # 平日: 作業内容をクリア、備考は在宅をデフォルト
            safe_set(row_cells[COL_CONTENT], '')
            safe_set(row_cells[COL_NOTE],    '在宅')

        day_idx += 1

    # シート名も更新
    if ws.title and '月度' in ws.title:
        try:
            ws.title = f'作業報告書_{y}年{m}月度'
        except Exception:
            pass

    wb.save(str(new_path))
    return {
        'ok':       True,
        'path':     str(new_path),
        'filename': new_filename,
        'year':     next_year,
        'month':    next_month,
    }


# ===== 単体テスト =====

if __name__ == '__main__':
    print('=== list_reports ===')
    reports = list_reports()
    for r in reports:
        print(f"  {r['year']}-{r['month']} | {r['status']:10} | {r['filename']}")

    if reports:
        latest = reports[-1]
        print(f"\n=== read_report({latest['year']}, {latest['month']}) ===")
        data = read_report(latest['year'], latest['month'])
        if data:
            print(f"  meta: {data['meta']}")
            print(f"  days: {len(data['days'])} 件")
            for d in data['days'][:3]:
                print(f"    {d}")
        else:
            print('  読み込み失敗')
